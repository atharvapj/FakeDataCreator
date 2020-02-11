__author__ = "Rajib Ahmed"
__version__ = "1.0"

print("Importing Modules")
import math
print("Imported Math")
from faker import Faker
from faker.providers import internet
from faker_e164.providers import E164Provider
import faker.providers.credit_card
fake = Faker()
fake.add_provider(E164Provider)
print("Imported Faker")
from openpyxl import load_workbook
from openpyxl import *
print("Imported OpenPYXL")
import names
print("Imported Names")
import os
print("Imported OS")
import string
print("Imported String")
from concurrent.futures import ThreadPoolExecutor, Future
print("Imported Concurrent Futures")
import multiprocessing
print("Imported MultiProcessing")
from datetime import datetime
import datetime
print("Imported DateTime")
import random
print("Imported Random")
import time
print("Imported Time")
import binascii
print("Imported binascii")
print("Modules Imported Successfully!")
print("Loading Sheet into Memory")
wb = load_workbook("samplesheet.xlsx")##THIS IS WHERE YOUR SHEET WILL GO
ws = wb.active
print("Sheet Loaded")

def saveWB():
    wb.save("samplesheet.xlsx")##replace sheetname here too!
def fName(gender,x,pos):
    cell= "B"+str(pos)
    fname = names.get_first_name(gender=gender)
    ws[cell] = fname
    print("Set cell " + cell + " First Name to " + fname)
def lName(gender,x,pos):
    cell= "C"+str(pos)
    lname = names.get_last_name()
    ws[cell] = lname
    print("Set cell " + cell + " Last Name to " + lname)
    

def doID(x):
    List1 = []
    j = 2
    while True:
        cellc = "A" + str(j)
        if ws[cellc].value == None:
            break
        else:
            List1.append(ws[cellc].value)
            j+=1
    for i in range(2,x):

        cell = "A" + str(i)
        ident = (((str(binascii.hexlify(os.urandom(6)))).replace("'","")).replace("b",""))

        while True:
            if ident in List1:
                ident = (((str(binascii.hexlify(os.urandom(6)))).replace("'","")).replace("b",""))
            else:
                break
        
            
        ws[cell] = ident
        print("Cell " + cell + " has been set to ID " + ident)
    
def Gender(x):
    genders = ['male','female']
    for i in range(2,x):
        g = random.choice(genders)
        cell = "D"+str(i)
        ws[cell] = g
        print("Set cell " + cell + " Gender to " + g)
        print("Attempting to set First Name")
        fName(g,x,i)
        print("Attempting to set Last Name")
        lName(g,x,i)

def dob(age,i):
    Cyear = int(datetime.datetime.now().year)
    dobyear = Cyear - age
    try:
        Dob =  datetime.datetime.strptime('{} {}'.format(random.randint(1, 366), dobyear), '%j %Y')
        print("Got Date Of Birth")
    except ValueError:
        print("Whoops we got a leap year by accident")

    cell = "E" + str(i)
    Dob = str(Dob)
    DobList = Dob.split(" ")
    Dob = DobList[0]
    ws[cell] = Dob
    print("Set cell " + cell + " Date Of Birth to " + str(Dob))
    
    

def Age(x):
    for i in range(2,x):
        cell = "G" + str(i)
        age = random.randint(13,99)
        ws[cell] = age
        print("Set cell " + cell + " Age to " + str(age))
        print("Attempting to Set Date Of Birth")
        dob(age,i)

def Address(x):
    for i in range(2,x):
        cell  = "F" + str(i)
        addr = fake.address()
        ws[cell] = fake.address()
        print("Set cell " + cell + " Address to " + addr)

def Phone(x):
    for i in range(2,x):
        cell = "H" + str(i)
        pnum = fake.e164(region_code="US", valid=True, possible=True)
        ws[cell] = pnum
        print("Set cell " + cell + " Phone Number to " + str(pnum))
def CardDet(x):
    for i in range(2,x):
        cell1 = "I" + str(i)
        cell2 = "J" + str(i)
        cell3 = "K" + str(i)
        cnum =fake.credit_card_number(card_type=None)
        cprov =fake.credit_card_provider(card_type=None)
        cccv = fake.credit_card_security_code(card_type=None)
        ws[cell1] = str(cnum)
        ws[cell2] = str(cprov)
        ws[cell3] = str(cccv)
        print("Set cell " + cell + " Card Number to " + str(cnum))
        print("Set cell " + cell + " Card Provider " + str(cprov))
        print("Set cell " + cell + " Card CCV " + str(cccv))

def start(x):
    print("Starting in 5 seconds")
    #time.sleep(5)
    x+=1
    print("Checking Open State")
    try:
        wb.save("Data3.xlsx")
    except:
        print("Please Close All instances of the sheet")
        return()
    print("Starting")
    pool = ThreadPoolExecutor(max_workers=100)
    pool.submit(doID,x)
    pool.submit(Gender,x)
    pool.submit(Age,x)
    pool.submit(Address,x)
    pool.submit(Phone,x)
    pool.submit(CardDet,x)
    pool.shutdown()
    print("Saving")
    wb.save("Data3.xlsx")
    print("Done!")

#These are just pools so we can multiprocess them. Not sure if splitting pools into pools makes a difference

def doP1(x):
    pool1 = ThreadPoolExecutor(max_workers=999)
    pool1.submit(doID,x)
    pool1.shutdown()
def doP2(x):
    pool2 = ThreadPoolExecutor(max_workers=999)
    pool2.submit(Address,x)
    pool2.shutdown()
def doP3(x):
    pool3 = ThreadPoolExecutor(max_workers=999)
    pool3.submit(Gender,x)
    pool3.shutdown()
def doP4(x):
    pool4 = ThreadPoolExecutor(max_workers=999)
    pool4.submit(Age,x)
    pool4.shutdown()
def doP5(x):
    pool5 = ThreadPoolExecutor(max_workers=999)
    pool5.submit(Phone,x)
    pool5.shutdown()
def doP6(x):
    pool6 = ThreadPoolExecutor(max_workers=999)
    pool6.submit(CardDet,x)
    pool6.shutdown()
def doP7(x):
    pool7 = ThreadPoolExecutor(max_workers=999)
    pool7.submit(doP1,x)
    pool7.submit(doP2,x)
    pool7.submit(doP3,x)
    pool7.submit(doP4,x)
    pool7.submit(doP5,x)
    pool7.submit(doP6,x)
    pool7.shutdown()
    saveWB()

doP7(int(input("Enter the cell number youd like to go to")))

        

    
